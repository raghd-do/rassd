from openpyxl import *
import os
import itertools

tests = os.listdir("test_templates")
database = load_workbook("database\قاعدة الفصل.xlsx")

def work_in_hefeth(database, student_name, student_id, student_grades):
    hefeth_sheet = database["حفظ"]
    found = False

    for row in hefeth_sheet.iter_rows(min_row=2, max_row=len(list(hefeth_sheet.rows))):
        if row[3].value == student_name and row[4].value == student_id:
            print("we found her in hefeth :D")
            found = True

            print("now we'r copying her marks")
            # قائمة بالخلايا الي راح نلصق فيهم درجات الطالبة بقاعدة الفصل
            target_cells = [hefeth_sheet.cell(row=row[1].row, column=col_index) for col_index in range(11, 61)]

            # نسخ الدرجات
            for grd_cell, tgt_cell in zip(student_grades[0], target_cells):
                print(grd_cell, tgt_cell)
                if grd_cell.value is not None:
                    tgt_cell.value = grd_cell.value
                else:
                    break  # Stop when the grade cell is empty
            
            print("copying is done :)")
            # حفظ التغييرات بقاعدة الفصل
            database.save(
                filename="database\قاعدة الفصل.xlsx")

    return found

def work_in_t3ahod(database, student_name, student_id, student_grades):
    t3ahod_sheet = database["تعاهد"]
    found = False

    for row in t3ahod_sheet.iter_rows(min_row=2, max_row=len(list(t3ahod_sheet.rows))):
        if row[3].value == student_name and row[4].value == student_id:
            print("we found her in t3ahod :D")
            found = True

            print("now we'r copying her marks")
            # قائمة بالخلايا الي راح نلصق فيهم درجات الطالبة بقاعدة الفصل
            target_cells = [t3ahod_sheet.cell(row=row[1].row, column=col_index) for col_index in range(11, 61)]

            # نسخ الدرجات
            for grd_cell, tgt_cell in zip(student_grades[0], target_cells):
                print(grd_cell, tgt_cell)
                if grd_cell.value is not None:
                    tgt_cell.value = grd_cell.value
                else:
                    break  # Stop when the grade cell is empty
            
            print("copying is done :)")
            # حفظ التغييرات بقاعدة الفصل
            database.save(
                filename="database\قاعدة الفصل.xlsx")

    return found

for test in tests:
    student_test = load_workbook("test_templates/" + test, data_only=True)
    sheet = student_test.active
    student_name = sheet["C4"].value
    student_track = sheet.title
    student_id = sheet["G4"].value if student_track == '6' else sheet["I4"].value
    student_grades = sheet["H1:AK1"] if student_track == '6' else sheet["J1:BH1"]

    print("Loking for:", student_name, student_id)

    found = work_in_hefeth(database, student_name, student_id, student_grades) # search and copy to hefeth
    if found:
        # move to done file
        isExist = os.path.exists("done")
        if not isExist:
            os.makedirs("done")

        os.rename(f"test_templates\{test}", f"done\{test}")

    else:
        isExist = os.path.exists("not_found")
        if not isExist:
            os.makedirs("not_found")

        print("Sorry we couldn't find her :(")
        os.rename(f"test_templates\{test}", f"not_found\{test}")
